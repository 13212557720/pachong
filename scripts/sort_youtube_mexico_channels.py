#!/usr/bin/env python3
"""Sort YouTube Mexico channel workbook by Mexico flag and email availability."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_INPUT = Path(
    "/Users/biantongchuangchuanmei/Desktop/pachong/"
    "output_youtube_mexico_final_40k_10k_500k_relaxed/"
    "youtube_mexico_10k_500k_channels.xlsx"
)

MEXICO_COLUMN = "是否严格墨西哥"
EMAIL_COLUMN = "邮箱"


def normalize_text(value: Any) -> str:
    """Return a trimmed text value; empty cells become an empty string."""
    if value is None:
        return ""
    return str(value).strip()


def is_email_present(value: Any) -> bool:
    """Treat any non-empty email cell as having an email."""
    return bool(normalize_text(value))


def build_default_output_path(input_path: Path) -> Path:
    """Create an output path next to the input workbook."""
    return input_path.with_name(f"{input_path.stem}_sorted{input_path.suffix}")


def snapshot_cell(cell: Any) -> dict[str, Any]:
    """Capture value and style information so sorting keeps the sheet appearance."""
    return {
        "value": cell.value,
        "data_type": cell.data_type,
        "style": copy(cell._style),
        "number_format": cell.number_format,
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
        "hyperlink": copy(cell.hyperlink),
        "comment": copy(cell.comment),
    }


def restore_cell(cell: Any, data: dict[str, Any]) -> None:
    """Write a captured cell back into a worksheet."""
    cell.value = data["value"]
    cell.data_type = data["data_type"]
    cell._style = copy(data["style"])
    cell.number_format = data["number_format"]
    cell.font = copy(data["font"])
    cell.fill = copy(data["fill"])
    cell.border = copy(data["border"])
    cell.alignment = copy(data["alignment"])
    cell.protection = copy(data["protection"])
    cell.hyperlink = copy(data["hyperlink"])
    cell.comment = copy(data["comment"])


def find_column_index(headers: list[Any], column_name: str) -> int:
    """Find a header column by exact name after trimming whitespace."""
    normalized_headers = [normalize_text(header) for header in headers]
    if column_name not in normalized_headers:
        raise ValueError(
            f"未找到列：{column_name}。当前表头：{', '.join(normalized_headers)}"
        )
    return normalized_headers.index(column_name)


def sort_workbook(input_path: Path, output_path: Path, sheet_name: str | None = None) -> dict[str, int]:
    """Sort workbook rows and save to a new output path."""
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")
    if input_path.resolve() == output_path.resolve():
        raise ValueError("输出文件不能和输入文件相同，避免误覆盖原表。")

    workbook = load_workbook(input_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    if worksheet.max_row < 2:
        workbook.save(output_path)
        return {"total": 0, "mexico_with_email": 0, "mexico_without_email": 0, "other": 0}

    headers = [cell.value for cell in worksheet[1]]
    mexico_index = find_column_index(headers, MEXICO_COLUMN)
    email_index = find_column_index(headers, EMAIL_COLUMN)

    rows: list[tuple[int, list[dict[str, Any]]]] = []
    stats = {"total": 0, "mexico_with_email": 0, "mexico_without_email": 0, "other": 0}

    for original_order, row in enumerate(
        worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column)
    ):
        captured_row = [snapshot_cell(cell) for cell in row]
        is_mexico = normalize_text(captured_row[mexico_index]["value"]) == "是"
        has_email = is_email_present(captured_row[email_index]["value"])

        stats["total"] += 1
        if is_mexico and has_email:
            stats["mexico_with_email"] += 1
        elif is_mexico:
            stats["mexico_without_email"] += 1
        else:
            stats["other"] += 1

        rows.append((original_order, captured_row))

    def sort_key(item: tuple[int, list[dict[str, Any]]]) -> tuple[int, int, int]:
        original_order, captured_row = item
        is_mexico = normalize_text(captured_row[mexico_index]["value"]) == "是"
        has_email = is_email_present(captured_row[email_index]["value"])
        return (
            0 if is_mexico else 1,
            0 if has_email else 1,
            original_order,
        )

    sorted_rows = sorted(rows, key=sort_key)

    for target_row_number, (_, captured_row) in enumerate(sorted_rows, start=2):
        for target_col_number, captured_cell in enumerate(captured_row, start=1):
            restore_cell(worksheet.cell(row=target_row_number, column=target_col_number), captured_cell)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "按“是否严格墨西哥=是”优先，再按“邮箱非空”优先排序 YouTube 频道 Excel。"
        )
    )
    parser.add_argument(
        "input",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT,
        help="输入 Excel 文件路径。默认处理当前指定的 YouTube Mexico 表格。",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="输出 Excel 文件路径。默认在输入文件旁边生成 *_sorted.xlsx。",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="要处理的工作表名称。默认处理活动工作表。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else build_default_output_path(input_path).resolve()
    )

    stats = sort_workbook(input_path=input_path, output_path=output_path, sheet_name=args.sheet)

    print(f"处理完成：{output_path}")
    print(f"总数据行数：{stats['total']}")
    print(f"墨西哥有邮箱：{stats['mexico_with_email']}")
    print(f"墨西哥无邮箱：{stats['mexico_without_email']}")
    print(f"其他：{stats['other']}")


if __name__ == "__main__":
    main()
